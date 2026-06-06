"""
================================================================================
  ARCHIVO   : etl_landing.py
  PROYECTO  : Sistema de Gestión de Inventario — Fertilisa S.A.
  MÓDULO    : ETL Capa 0 — Excel → fertiliza_landing
  VERSIÓN   : 3.0 — Junio 2026

  ARCHIVOS PROCESADOS:
    Resumen.xlsx
      - Hoja "Fertilisa"        → lnd_fertilisa_raw        (211 SKUs)
      - Hoja "Manvert, Jiffy"   → lnd_manvert_jiffy_raw    (26 SKUs)

    ARCHIVO FERTILISA-INVESTIGACION.xlsx
      - Hoja "LISTA DE PRODUCTOS"   → lnd_lista_productos_raw (712 productos)
      - Hoja "EJEMPLO DATOS KARDEX" → lnd_kardex_raw          (731 movimientos)

  ESTRUCTURA REAL VERIFICADA:
    Fertilisa:
      Col B(1): código (apóstrofe líder), Col C(2): nombre, Col D(3): físico,
      Col E(4): sistema, Col F(5): diferencia, Col G(6): observación
      Secciones en Col B: "Granular - Solubles", "Liquidos", "Material para Invernaderos..."

    Manvert/Jiffy:
      Col A(0): lote, Col B(1): código, Col C(2): nombre, Col D(3): estado lote,
      Col E(4): físico, Col F(5): sistema, Col G(6): diferencia, Col H(7): observación,
      Col I(8): fecha vencimiento, Col J(9): #REF! IGNORAR, Col L(11): almacenaje

    Lista de Productos:
      Col A(0): código, Col B(1): descripción, Col C(2): precio1,
      Col D(3): precio2, Col E(4): precio3, Col F(5): existencia

    Kardex:
      Col A(0): tipo, Col B(1): fecha, Col C(2): documento,
      Col D(3): cliente/proveedor, Col E(4): costo, Col F(5): entradas,
      Col G(6): precio, Col H(7): salidas, Col I(8): bodega, Col J(9): cantidad_toma

  USO:
    python etl_landing.py --input-dir /ruta/carpeta [--log-dir ./logs] [--usuario nombre]

  DEPENDENCIAS:
    pip install openpyxl psycopg2-binary python-dotenv
================================================================================
"""

import argparse
import hashlib
import logging
import os
import re
import sys
from datetime import datetime, date
from pathlib import Path
from typing import Optional, Tuple

import openpyxl
import psycopg2
import psycopg2.extras
from dotenv import load_dotenv


# ─────────────────────────────────────────────────────────────────────────────
#  CONSTANTES
# ─────────────────────────────────────────────────────────────────────────────

SCRIPT_VERSION = "3.0.0"

HOJAS_CONOCIDAS = {
    "Fertilisa":             "fertilisa",
    "Manvert, Jiffy":        "manvert_jiffy",
    "Manvert,Jiffy":         "manvert_jiffy",
    "Manvert Jiffy":         "manvert_jiffy",
    "LISTA DE PRODUCTOS":    "lista_productos",
    "EJEMPLO DATOS KARDEX":  "kardex",
}

NOMBRES_SECCION = {
    "granular - solubles":                           "Granular - Solubles",
    "granular-solubles":                             "Granular - Solubles",
    "liquidos":                                      "Liquidos",
    "líquidos":                                      "Liquidos",
    "material para invernaderos - sustrato y otros": "Material para Invernaderos - Sustrato y Otros",
    "material para invernaderos":                    "Material para Invernaderos - Sustrato y Otros",
}

VALORES_ENCABEZADO = {"codigo", "código", "cod", "code", "tipo"}

TIPOS_KARDEX_RECONOCIDOS = {
    "Entrada Inventario-SI",
    "Salida Inventario-SI",
    "Toma Física",
    "Venta-CONSUMIDOR FINAL-SI",
    "Venta-NOTA DE CRÉDITO FISOFT-SI",
    "Entrada",
}


# ─────────────────────────────────────────────────────────────────────────────
#  ÍNDICES DE COLUMNAS
# ─────────────────────────────────────────────────────────────────────────────

class ColFertilisa:
    A          = 0
    CODIGO     = 1
    NOMBRE     = 2
    FISICO     = 3
    SISTEMA    = 4
    DIFERENCIA = 5
    OBSERVACION = 6

class ColManvert:
    LOTE        = 0
    CODIGO      = 1
    NOMBRE      = 2
    ESTADO      = 3
    FISICO      = 4
    SISTEMA     = 5
    DIFERENCIA  = 6
    OBSERVACION = 7
    FECHA_VENC  = 8
    # Col J (9) = #REF! IGNORAR SIEMPRE
    # Col K (10) = None IGNORAR
    ALMACENAJE  = 11

class ColListaProductos:
    CODIGO      = 0
    DESCRIPCION = 1
    PRECIO1     = 2
    PRECIO2     = 3
    PRECIO3     = 4
    EXISTENCIA  = 5

class ColKardex:
    TIPO         = 0
    FECHA        = 1
    DOCUMENTO    = 2
    CLIENTE_PROV = 3
    COSTO        = 4
    ENTRADAS     = 5
    PRECIO       = 6
    SALIDAS      = 7
    BODEGA       = 8
    CANTIDAD_TOMA = 9


# ─────────────────────────────────────────────────────────────────────────────
#  LOGGING
# ─────────────────────────────────────────────────────────────────────────────

def configurar_logging(log_dir: Optional[str]) -> logging.Logger:
    logger = logging.getLogger("etl_landing")
    logger.setLevel(logging.DEBUG)
    fmt = logging.Formatter(
        fmt="%(asctime)s | %(levelname)-8s | %(message)s",
        datefmt="%Y-%m-%d %H:%M:%S"
    )
    ch = logging.StreamHandler(sys.stdout)
    ch.setLevel(logging.INFO)
    ch.setFormatter(fmt)
    logger.addHandler(ch)

    if log_dir:
        Path(log_dir).mkdir(parents=True, exist_ok=True)
        ts = datetime.now().strftime("%Y%m%d_%H%M%S")
        log_path = Path(log_dir) / f"etl_landing_{ts}.log"
        fh = logging.FileHandler(log_path, encoding="utf-8")
        fh.setLevel(logging.DEBUG)
        fh.setFormatter(fmt)
        logger.addHandler(fh)
        logger.info(f"Log de archivo: {log_path}")

    return logger


# ─────────────────────────────────────────────────────────────────────────────
#  CONEXIÓN A BASE DE DATOS
# ─────────────────────────────────────────────────────────────────────────────

def conectar_db(logger: logging.Logger) -> psycopg2.extensions.connection:
    load_dotenv()
    params = {
        "host":     os.getenv("DB_HOST", "localhost"),
        "port":     os.getenv("DB_PORT", "5432"),
        "dbname":   os.getenv("DB_NAME", "fertilisa"),
        "user":     os.getenv("DB_USER", "postgres"),
        "password": os.getenv("DB_PASSWORD", ""),
        "sslmode":  os.getenv("DB_SSLMODE", "prefer"),
    }
    try:
        conn = psycopg2.connect(**params)
        conn.autocommit = False
        logger.info(
            f"Conectado a PostgreSQL: "
            f"{params['host']}:{params['port']}/{params['dbname']}"
        )
        return conn
    except psycopg2.OperationalError as e:
        logger.critical(f"ERROR DE CONEXIÓN: {e}")
        logger.critical("Proceso detenido. Verificar credenciales y servidor.")
        sys.exit(1)


# ─────────────────────────────────────────────────────────────────────────────
#  UTILIDADES DE LIMPIEZA
# ─────────────────────────────────────────────────────────────────────────────

def celda_a_str(valor) -> Optional[str]:
    if valor is None:
        return None
    if isinstance(valor, (datetime, date)):
        return valor.strftime("%Y-%m-%d")
    texto = str(valor).strip()
    return texto if texto else None

def limpiar_codigo(valor_raw: Optional[str]) -> Optional[str]:
    if not valor_raw:
        return None
    limpio = str(valor_raw).lstrip("'").strip()
    return limpio if limpio else None

def limpiar_nombre(valor_raw: Optional[str]) -> Optional[str]:
    if not valor_raw:
        return None
    limpio = re.sub(r" +", " ", str(valor_raw).strip())
    return limpio if limpio else None

def es_numerico(valor_raw: Optional[str]) -> bool:
    if not valor_raw:
        return False
    try:
        float(str(valor_raw).replace(",", "").strip())
        return True
    except ValueError:
        return False

def es_fecha_parseable(valor_raw: Optional[str]) -> bool:
    if not valor_raw:
        return True
    if valor_raw.upper() in ("N/A", "NA", "NONE", ""):
        return True
    for fmt in ("%Y-%m-%d", "%Y-%m-%dT%H:%M:%S", "%d/%m/%Y", "%m/%d/%Y", "%d-%m-%Y"):
        try:
            datetime.strptime(valor_raw[:10], fmt)
            return True
        except ValueError:
            continue
    return False

def detectar_seccion_fertilisa(col_b_raw: Optional[str]) -> Optional[str]:
    if not col_b_raw:
        return None
    return NOMBRES_SECCION.get(str(col_b_raw).strip().lower())

def es_encabezado_columnas(col_b_raw: Optional[str]) -> bool:
    if not col_b_raw:
        return False
    limpio = limpiar_codigo(col_b_raw)
    return bool(limpio) and limpio.lower() in VALORES_ENCABEZADO

def es_fila_vacia(row: tuple, indices: list) -> bool:
    return not any(
        celda_a_str(row[i] if len(row) > i else None)
        for i in indices
    )

def extraer_celda(row: tuple, idx: int) -> Optional[str]:
    return celda_a_str(row[idx] if len(row) > idx else None)

def calcular_flags_fertilisa(codigo_raw, nombre_raw, fisico_raw, sistema_raw) -> dict:
    tiene_codigo    = bool(limpiar_codigo(codigo_raw))
    tiene_nombre    = bool(limpiar_nombre(nombre_raw))
    fisico_num      = es_numerico(fisico_raw)
    sistema_num     = es_numerico(sistema_raw)
    return {
        "lnd_tiene_codigo":     tiene_codigo,
        "lnd_tiene_nombre":     tiene_nombre,
        "lnd_fisico_numerico":  fisico_num,
        "lnd_sistema_numerico": sistema_num,
        "lnd_listo_para_stg":   tiene_codigo and tiene_nombre and fisico_num,
    }

def calcular_flags_manvert(codigo_raw, nombre_raw, fisico_raw, sistema_raw,
                           fecha_venc_raw, almacenaje_raw) -> dict:
    tiene_codigo   = bool(limpiar_codigo(codigo_raw))
    tiene_nombre   = bool(limpiar_nombre(nombre_raw))
    fisico_num     = es_numerico(fisico_raw)
    sistema_num    = es_numerico(sistema_raw)
    fecha_parse    = es_fecha_parseable(fecha_venc_raw)
    almac_num      = es_numerico(almacenaje_raw)
    return {
        "lnd_tiene_codigo":          tiene_codigo,
        "lnd_tiene_nombre":          tiene_nombre,
        "lnd_fisico_numerico":       fisico_num,
        "lnd_sistema_numerico":      sistema_num,
        "lnd_fecha_venc_parseable":  fecha_parse,
        "lnd_almacenaje_numerico":   almac_num,
        "lnd_listo_para_stg":        tiene_codigo and tiene_nombre and fisico_num,
    }

def calcular_flags_lista_productos(codigo_raw, descripcion_raw,
                                   precio1_raw, existencia_raw) -> dict:
    tiene_codigo  = bool(limpiar_codigo(codigo_raw))
    tiene_desc    = bool(limpiar_nombre(descripcion_raw))
    precio1_num   = es_numerico(precio1_raw)
    exist_num     = es_numerico(existencia_raw)
    return {
        "lnd_tiene_codigo":         tiene_codigo,
        "lnd_tiene_descripcion":    tiene_desc,
        "lnd_precio1_numerico":     precio1_num,
        "lnd_existencia_numerica":  exist_num,
        "lnd_listo_para_stg":       tiene_codigo and tiene_desc,
    }

def calcular_flags_kardex(tipo_raw, fecha_raw, entradas_raw, salidas_raw) -> dict:
    tiene_tipo      = bool(tipo_raw and tipo_raw.strip())
    fecha_parse     = es_fecha_parseable(fecha_raw)
    entradas_num    = es_numerico(entradas_raw)
    salidas_num     = es_numerico(salidas_raw)
    tipo_reconocido = (tipo_raw or "").strip() in TIPOS_KARDEX_RECONOCIDOS
    return {
        "lnd_tiene_tipo":         tiene_tipo,
        "lnd_fecha_parseable":    fecha_parse,
        "lnd_entradas_numericas": entradas_num,
        "lnd_salidas_numericas":  salidas_num,
        "lnd_tipo_reconocido":    tipo_reconocido,
        "lnd_listo_para_stg":     tiene_tipo and fecha_parse and tipo_reconocido,
    }


# ─────────────────────────────────────────────────────────────────────────────
#  MD5
# ─────────────────────────────────────────────────────────────────────────────

def calcular_md5(ruta: Path) -> str:
    h = hashlib.md5()
    with open(ruta, "rb") as f:
        for chunk in iter(lambda: f.read(8192), b""):
            h.update(chunk)
    return h.hexdigest()


# ─────────────────────────────────────────────────────────────────────────────
#  OPERACIONES EN BASE DE DATOS
# ─────────────────────────────────────────────────────────────────────────────

def verificar_duplicado(cur, hash_md5: str, logger: logging.Logger) -> bool:
    cur.execute("""
        SELECT id_archivo_log, nombre_archivo, inicio_proceso
        FROM fertiliza_landing.lnd_archivo_log
        WHERE hash_md5 = %s AND estado = 'Procesado'
        LIMIT 1
    """, (hash_md5,))
    row = cur.fetchone()
    if row:
        logger.warning(
            f"DUPLICADO — MD5 {hash_md5} ya procesado "
            f"(id={row[0]}, archivo='{row[1]}', fecha={row[2]}). Abortando."
        )
        return True
    return False

def registrar_archivo_log(cur, nombre, ruta, hash_md5, tamanio, usuario) -> int:
    cur.execute("""
        INSERT INTO fertiliza_landing.lnd_archivo_log
            (nombre_archivo, ruta_origen, hash_md5, tamanio_bytes,
             estado, ejecutado_por, script_version, inicio_proceso)
        VALUES (%s,%s,%s,%s,'En Proceso',%s,%s,NOW())
        RETURNING id_archivo_log
    """, (nombre, ruta, hash_md5, tamanio, usuario, SCRIPT_VERSION))
    return cur.fetchone()[0]

def actualizar_archivo_log(cur, id_archivo, estado, total_filas,
                           filas_insertadas, mensaje=None):
    cur.execute("""
        UPDATE fertiliza_landing.lnd_archivo_log
        SET estado=%s, total_filas_raw=%s, filas_insertadas=%s,
            mensaje=%s, fin_proceso=NOW()
        WHERE id_archivo_log=%s
    """, (estado, total_filas, filas_insertadas, mensaje, id_archivo))

def registrar_carga(cur, id_archivo, hoja, periodo, usuario) -> int:
    cur.execute("""
        INSERT INTO fertiliza_landing.lnd_carga
            (id_archivo_log, hoja_procesada, periodo_datos,
             estado, ejecutado_por, inicio_carga)
        VALUES (%s,%s,%s,'Iniciado',%s,NOW())
        RETURNING id_carga
    """, (id_archivo, hoja, periodo, usuario))
    return cur.fetchone()[0]

def actualizar_carga(cur, id_carga, estado, leidas, vacias,
                     secciones, insertadas, mensaje=None):
    cur.execute("""
        UPDATE fertiliza_landing.lnd_carga
        SET estado=%s, filas_leidas=%s, filas_vacias=%s,
            filas_seccion=%s, filas_insertadas=%s,
            mensaje_error=%s, fin_carga=NOW()
        WHERE id_carga=%s
    """, (estado, leidas, vacias, secciones, insertadas, mensaje, id_carga))


# ─────────────────────────────────────────────────────────────────────────────
#  PROCESAMIENTO: HOJA FERTILISA
# ─────────────────────────────────────────────────────────────────────────────

SQL_INSERT_FERTILISA = """
    INSERT INTO fertiliza_landing.lnd_fertilisa_raw (
        id_carga, fila_excel, hoja_origen,
        col_a_raw, codigo_raw, nombre_raw,
        fisico_raw, sistema_raw, diferencia_raw, observacion_raw,
        categoria_raw, es_fila_seccion, es_fila_vacia,
        lnd_tiene_codigo, lnd_tiene_nombre,
        lnd_fisico_numerico, lnd_sistema_numerico,
        lnd_listo_para_stg, creado_en
    ) VALUES %s
"""

def procesar_hoja_fertilisa(ws, cur, id_carga: int, logger: logging.Logger) -> dict:
    cnt = {"leidas":0,"vacias":0,"secciones":0,"encabezados":0,"insertadas":0,"warnings":0}
    categoria_actual: Optional[str] = None
    buffer = []
    BATCH = 200

    for idx, row in enumerate(ws.iter_rows(min_row=4, values_only=True), start=4):
        cnt["leidas"] += 1

        col_a = extraer_celda(row, ColFertilisa.A)
        col_b = extraer_celda(row, ColFertilisa.CODIGO)
        col_c = extraer_celda(row, ColFertilisa.NOMBRE)
        col_d = extraer_celda(row, ColFertilisa.FISICO)
        col_e = extraer_celda(row, ColFertilisa.SISTEMA)
        col_f = extraer_celda(row, ColFertilisa.DIFERENCIA)
        col_g = extraer_celda(row, ColFertilisa.OBSERVACION)

        if es_fila_vacia(row, [ColFertilisa.CODIGO, ColFertilisa.NOMBRE,
                                ColFertilisa.FISICO, ColFertilisa.SISTEMA]):
            cnt["vacias"] += 1
            continue

        if es_encabezado_columnas(col_b):
            cnt["encabezados"] += 1
            continue

        nombre_sec = detectar_seccion_fertilisa(col_b)
        if nombre_sec:
            categoria_actual = nombre_sec
            cnt["secciones"] += 1
            logger.info(f"  F{idx}: sección → '{categoria_actual}'")
            continue

        flags = calcular_flags_fertilisa(col_b, col_c, col_d, col_e)
        cod_limpio = limpiar_codigo(col_b)

        if not flags["lnd_sistema_numerico"] and col_d:
            cnt["warnings"] += 1
            logger.warning(f"  F{idx} RV-04: sistema vacío (código='{cod_limpio}')")
        if flags["lnd_tiene_codigo"] and not flags["lnd_tiene_nombre"]:
            cnt["warnings"] += 1
            logger.warning(f"  F{idx} RV-02: código sin nombre ('{cod_limpio}')")
        elif flags["lnd_tiene_nombre"] and not flags["lnd_tiene_codigo"]:
            cnt["warnings"] += 1
            logger.warning(f"  F{idx} RV-01: nombre sin código")

        buffer.append((
            id_carga, idx, "Fertilisa",
            col_a, col_b, col_c, col_d, col_e, col_f,
            col_g if col_g and col_g.strip() else None,
            categoria_actual, False, False,
            flags["lnd_tiene_codigo"], flags["lnd_tiene_nombre"],
            flags["lnd_fisico_numerico"], flags["lnd_sistema_numerico"],
            flags["lnd_listo_para_stg"], datetime.now()
        ))
        cnt["insertadas"] += 1

        if len(buffer) >= BATCH:
            psycopg2.extras.execute_values(cur, SQL_INSERT_FERTILISA, buffer)
            buffer.clear()

    if buffer:
        psycopg2.extras.execute_values(cur, SQL_INSERT_FERTILISA, buffer)

    return cnt


# ─────────────────────────────────────────────────────────────────────────────
#  PROCESAMIENTO: HOJA MANVERT/JIFFY
# ─────────────────────────────────────────────────────────────────────────────

SQL_INSERT_MANVERT = """
    INSERT INTO fertiliza_landing.lnd_manvert_jiffy_raw (
        id_carga, fila_excel, hoja_origen,
        lote_raw, codigo_raw, nombre_raw, estado_lote_raw,
        fisico_raw, sistema_raw, diferencia_raw,
        observacion_raw, fecha_venc_raw, almacenaje_raw,
        es_fila_seccion, es_fila_vacia,
        lnd_tiene_codigo, lnd_tiene_nombre,
        lnd_fisico_numerico, lnd_sistema_numerico,
        lnd_fecha_venc_parseable, lnd_almacenaje_numerico,
        lnd_listo_para_stg, creado_en
    ) VALUES %s
"""

def procesar_hoja_manvert_jiffy(ws, cur, id_carga: int, logger: logging.Logger) -> dict:
    cnt = {"leidas":0,"vacias":0,"secciones":0,"encabezados":0,"insertadas":0,"warnings":0}
    buffer = []
    BATCH = 200

    for idx, row in enumerate(ws.iter_rows(min_row=4, values_only=True), start=4):
        cnt["leidas"] += 1

        lote  = extraer_celda(row, ColManvert.LOTE)
        cod   = extraer_celda(row, ColManvert.CODIGO)
        nom   = extraer_celda(row, ColManvert.NOMBRE)
        est   = extraer_celda(row, ColManvert.ESTADO)
        fis   = extraer_celda(row, ColManvert.FISICO)
        sis   = extraer_celda(row, ColManvert.SISTEMA)
        dif   = extraer_celda(row, ColManvert.DIFERENCIA)
        obs   = extraer_celda(row, ColManvert.OBSERVACION)
        fvenc = extraer_celda(row, ColManvert.FECHA_VENC)
        alm   = extraer_celda(row, ColManvert.ALMACENAJE)

        if es_fila_vacia(row, [ColManvert.CODIGO, ColManvert.NOMBRE,
                                ColManvert.FISICO, ColManvert.SISTEMA]):
            cnt["vacias"] += 1
            continue

        if es_encabezado_columnas(cod):
            cnt["encabezados"] += 1
            continue

        if est in ("N/A", "NA", ""):
            est = None
        elif est and es_numerico(est):
            est = str(int(float(est)))

        if fvenc in ("N/A", "NA", ""):
            fvenc = None

        flags = calcular_flags_manvert(cod, nom, fis, sis, fvenc, alm)
        cod_limpio = limpiar_codigo(cod)

        if not flags["lnd_sistema_numerico"] and fis:
            cnt["warnings"] += 1
            logger.warning(f"  F{idx} RV-04: sistema vacío (código='{cod_limpio}')")
        if not flags["lnd_fecha_venc_parseable"]:
            cnt["warnings"] += 1
            logger.warning(f"  F{idx}: fecha_venc '{fvenc}' no parseable")
        if flags["lnd_tiene_codigo"] and not flags["lnd_tiene_nombre"]:
            cnt["warnings"] += 1
            logger.warning(f"  F{idx} RV-02: código sin nombre")
        elif flags["lnd_tiene_nombre"] and not flags["lnd_tiene_codigo"]:
            cnt["warnings"] += 1
            logger.warning(f"  F{idx} RV-01: nombre sin código")

        buffer.append((
            id_carga, idx, "Manvert_Jiffy",
            lote, cod, nom, est, fis, sis, dif, obs, fvenc, alm,
            False, False,
            flags["lnd_tiene_codigo"], flags["lnd_tiene_nombre"],
            flags["lnd_fisico_numerico"], flags["lnd_sistema_numerico"],
            flags["lnd_fecha_venc_parseable"], flags["lnd_almacenaje_numerico"],
            flags["lnd_listo_para_stg"], datetime.now()
        ))
        cnt["insertadas"] += 1

        if len(buffer) >= BATCH:
            psycopg2.extras.execute_values(cur, SQL_INSERT_MANVERT, buffer)
            buffer.clear()

    if buffer:
        psycopg2.extras.execute_values(cur, SQL_INSERT_MANVERT, buffer)

    return cnt


# ─────────────────────────────────────────────────────────────────────────────
#  PROCESAMIENTO: HOJA LISTA DE PRODUCTOS
# ─────────────────────────────────────────────────────────────────────────────

SQL_INSERT_LISTA_PRODUCTOS = """
    INSERT INTO fertiliza_landing.lnd_lista_productos_raw (
        id_carga, fila_excel, hoja_origen,
        codigo_raw, descripcion_raw,
        precio1_raw, precio2_raw, precio3_raw, existencia_raw,
        lnd_tiene_codigo, lnd_tiene_descripcion,
        lnd_precio1_numerico, lnd_existencia_numerica,
        lnd_listo_para_stg, creado_en
    ) VALUES %s
"""

def procesar_hoja_lista_productos(ws, cur, id_carga: int, logger: logging.Logger) -> dict:
    cnt = {"leidas":0,"vacias":0,"secciones":0,"encabezados":0,"insertadas":0,"warnings":0}
    buffer = []
    BATCH = 200

    for idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        cnt["leidas"] += 1

        cod   = extraer_celda(row, ColListaProductos.CODIGO)
        desc  = extraer_celda(row, ColListaProductos.DESCRIPCION)
        p1    = extraer_celda(row, ColListaProductos.PRECIO1)
        p2    = extraer_celda(row, ColListaProductos.PRECIO2)
        p3    = extraer_celda(row, ColListaProductos.PRECIO3)
        exist = extraer_celda(row, ColListaProductos.EXISTENCIA)

        if es_fila_vacia(row, [ColListaProductos.CODIGO, ColListaProductos.DESCRIPCION]):
            cnt["vacias"] += 1
            continue

        if es_encabezado_columnas(cod):
            cnt["encabezados"] += 1
            continue

        flags = calcular_flags_lista_productos(cod, desc, p1, exist)
        cod_limpio = limpiar_codigo(cod)

        if not flags["lnd_precio1_numerico"] or (p1 and float(p1 or 0) == 0):
            if p1 == "0" or p1 is None:
                cnt["warnings"] += 1
                logger.debug(f"  F{idx}: precio en cero (código='{cod_limpio}')")

        buffer.append((
            id_carga, idx, "LISTA DE PRODUCTOS",
            cod, desc, p1, p2, p3, exist,
            flags["lnd_tiene_codigo"], flags["lnd_tiene_descripcion"],
            flags["lnd_precio1_numerico"], flags["lnd_existencia_numerica"],
            flags["lnd_listo_para_stg"], datetime.now()
        ))
        cnt["insertadas"] += 1

        if len(buffer) >= BATCH:
            psycopg2.extras.execute_values(cur, SQL_INSERT_LISTA_PRODUCTOS, buffer)
            buffer.clear()

    if buffer:
        psycopg2.extras.execute_values(cur, SQL_INSERT_LISTA_PRODUCTOS, buffer)

    return cnt


# ─────────────────────────────────────────────────────────────────────────────
#  PROCESAMIENTO: HOJA EJEMPLO DATOS KARDEX
# ─────────────────────────────────────────────────────────────────────────────

SQL_INSERT_KARDEX = """
    INSERT INTO fertiliza_landing.lnd_kardex_raw (
        id_carga, fila_excel, hoja_origen,
        tipo_raw, fecha_raw, documento_raw, cliente_prov_raw,
        costo_raw, entradas_raw, precio_raw, salidas_raw,
        bodega_raw, cantidad_toma_raw,
        lnd_tiene_tipo, lnd_fecha_parseable,
        lnd_entradas_numericas, lnd_salidas_numericas,
        lnd_tipo_reconocido, lnd_listo_para_stg, creado_en
    ) VALUES %s
"""

def procesar_hoja_kardex(ws, cur, id_carga: int, logger: logging.Logger) -> dict:
    cnt = {"leidas":0,"vacias":0,"secciones":0,"encabezados":0,"insertadas":0,"warnings":0}
    buffer = []
    BATCH = 200

    for idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        cnt["leidas"] += 1

        tipo   = extraer_celda(row, ColKardex.TIPO)
        fecha  = extraer_celda(row, ColKardex.FECHA)
        doc    = extraer_celda(row, ColKardex.DOCUMENTO)
        cli    = extraer_celda(row, ColKardex.CLIENTE_PROV)
        costo  = extraer_celda(row, ColKardex.COSTO)
        entr   = extraer_celda(row, ColKardex.ENTRADAS)
        precio = extraer_celda(row, ColKardex.PRECIO)
        sal    = extraer_celda(row, ColKardex.SALIDAS)
        bodega = extraer_celda(row, ColKardex.BODEGA)
        cant   = extraer_celda(row, ColKardex.CANTIDAD_TOMA)

        if es_fila_vacia(row, [ColKardex.TIPO, ColKardex.FECHA]):
            cnt["vacias"] += 1
            continue

        if tipo and tipo.strip().lower() == "tipo":
            cnt["encabezados"] += 1
            continue

        doc = limpiar_codigo(doc) if doc else None
        flags = calcular_flags_kardex(tipo, fecha, entr, sal)

        if flags["lnd_tiene_tipo"] and not flags["lnd_tipo_reconocido"]:
            cnt["warnings"] += 1
            logger.warning(
                f"  F{idx} RV-14: tipo de movimiento no reconocido: '{tipo}'"
            )

        if not flags["lnd_fecha_parseable"]:
            cnt["warnings"] += 1
            logger.warning(f"  F{idx}: fecha '{fecha}' no parseable")

        buffer.append((
            id_carga, idx, "EJEMPLO DATOS KARDEX",
            tipo, fecha, doc, cli, costo, entr, precio, sal, bodega, cant,
            flags["lnd_tiene_tipo"], flags["lnd_fecha_parseable"],
            flags["lnd_entradas_numericas"], flags["lnd_salidas_numericas"],
            flags["lnd_tipo_reconocido"], flags["lnd_listo_para_stg"],
            datetime.now()
        ))
        cnt["insertadas"] += 1

        if len(buffer) >= BATCH:
            psycopg2.extras.execute_values(cur, SQL_INSERT_KARDEX, buffer)
            buffer.clear()

    if buffer:
        psycopg2.extras.execute_values(cur, SQL_INSERT_KARDEX, buffer)

    return cnt


# ─────────────────────────────────────────────────────────────────────────────
#  CHECKLIST PRE-CARGA
# ─────────────────────────────────────────────────────────────────────────────

def checklist_pre_carga(wb, nombre_archivo: str,
                         logger: logging.Logger) -> Tuple[bool, list]:
    hojas_a_procesar = []
    procesados = set()

    for nombre_hoja, tipo in HOJAS_CONOCIDAS.items():
        if nombre_hoja in wb.sheetnames and tipo not in procesados:
            hojas_a_procesar.append((nombre_hoja, tipo))
            procesados.add(tipo)

    if not hojas_a_procesar:
        logger.warning(
            f"  [{nombre_archivo}] CHECK-4 FAIL: ninguna hoja conocida. "
            f"Disponibles: {wb.sheetnames}"
        )
        return False, []

    for hoja in wb.sheetnames:
        if hoja not in HOJAS_CONOCIDAS:
            logger.warning(f"  [{nombre_archivo}] Hoja desconocida ignorada: '{hoja}'")

    for nombre_esperado in HOJAS_CONOCIDAS:
        if nombre_esperado not in wb.sheetnames:
            logger.info(
                f"  [{nombre_archivo}] Hoja '{nombre_esperado}' no encontrada. "
                f"Se continúa con las disponibles."
            )

    if "Fertilisa" in wb.sheetnames:
        ws = wb["Fertilisa"]
        total = con_cod = 0
        for row in ws.iter_rows(min_row=4, values_only=True):
            col_b = extraer_celda(row, ColFertilisa.CODIGO)
            if (not es_fila_vacia(row, [ColFertilisa.CODIGO, ColFertilisa.NOMBRE,
                                        ColFertilisa.FISICO, ColFertilisa.SISTEMA])
                    and not es_encabezado_columnas(col_b)
                    and not detectar_seccion_fertilisa(col_b)):
                total += 1
                if limpiar_codigo(col_b):
                    con_cod += 1
        if total > 0:
            pct = (con_cod / total) * 100
            if pct < 80:
                logger.warning(f"  [{nombre_archivo}] CHECK-6 WARNING: {pct:.1f}% con código")
            else:
                logger.info(f"  [{nombre_archivo}] CHECK-6 OK: {pct:.1f}% de filas tienen código.")

    return True, hojas_a_procesar


# ─────────────────────────────────────────────────────────────────────────────
#  PROCESAMIENTO DE UN ARCHIVO
# ─────────────────────────────────────────────────────────────────────────────

PROCESADORES = {
    "fertilisa":       procesar_hoja_fertilisa,
    "manvert_jiffy":   procesar_hoja_manvert_jiffy,
    "lista_productos": procesar_hoja_lista_productos,
    "kardex":          procesar_hoja_kardex,
}

def procesar_archivo(ruta: Path, conn, logger: logging.Logger, usuario: str) -> bool:
    logger.info(f"{'─' * 60}")
    logger.info(f"Archivo: {ruta.name}")

    tamanio  = ruta.stat().st_size
    hash_md5 = calcular_md5(ruta)
    logger.info(f"  MD5: {hash_md5} | Tamaño: {tamanio:,} bytes")

    cur = conn.cursor()

    if verificar_duplicado(cur, hash_md5, logger):
        cur.close()
        return False

    id_archivo = registrar_archivo_log(
        cur, ruta.name, str(ruta.resolve()), hash_md5, tamanio, usuario
    )
    conn.commit()
    logger.info(f"  Registrado en lnd_archivo_log (id={id_archivo})")

    try:
        wb = openpyxl.load_workbook(ruta, read_only=False, data_only=True)
    except Exception as e:
        logger.error(f"  No se pudo abrir el Excel: {e}")
        actualizar_archivo_log(cur, id_archivo, "Error", 0, 0, str(e))
        conn.commit()
        cur.close()
        return False

    puede_continuar, hojas_a_procesar = checklist_pre_carga(wb, ruta.name, logger)
    if not puede_continuar:
        actualizar_archivo_log(cur, id_archivo, "Error", 0, 0,
                               "Checklist pre-carga fallido: hojas no reconocidas")
        conn.commit()
        wb.close()
        cur.close()
        return False

    periodo = datetime.now().strftime("%Y-%m")
    total_filas_global  = 0
    total_insert_global = 0
    alguna_ok           = False

    for nombre_hoja, tipo_hoja in hojas_a_procesar:
        logger.info(f"  → Procesando hoja '{nombre_hoja}'")
        ws = wb[nombre_hoja]

        id_carga = registrar_carga(cur, id_archivo, nombre_hoja, periodo, usuario)
        conn.commit()

        try:
            procesador_func = PROCESADORES[tipo_hoja]
            cnt = procesador_func(ws, cur, id_carga, logger)

            actualizar_carga(
                cur, id_carga, "Completado",
                cnt["leidas"], cnt["vacias"],
                cnt.get("secciones", 0), cnt["insertadas"]
            )
            conn.commit()

            total_filas_global  += cnt["leidas"]
            total_insert_global += cnt["insertadas"]
            alguna_ok = True

            logger.info(
                f"    ✓ Leídas:{cnt['leidas']} | Vacías:{cnt['vacias']} | "
                f"Secciones:{cnt.get('secciones', 0)} | Encabezados:{cnt.get('encabezados', 0)} | "
                f"Insertadas:{cnt['insertadas']} | Warnings:{cnt['warnings']}"
            )

        except Exception as e:
            conn.rollback()
            logger.error(f"    ERROR en hoja '{nombre_hoja}': {e}")
            try:
                actualizar_carga(cur, id_carga, "Error", 0, 0, 0, 0, str(e))
                conn.commit()
            except Exception:
                pass

    wb.close()

    estado_final = "Procesado" if alguna_ok else "Error"
    actualizar_archivo_log(
        cur, id_archivo, estado_final,
        total_filas_global, total_insert_global
    )
    conn.commit()
    cur.close()

    logger.info(
        f"  {'✓' if alguna_ok else '✗'} Estado: {estado_final} | "
        f"Total insertadas: {total_insert_global}"
    )
    return alguna_ok


# ─────────────────────────────────────────────────────────────────────────────
#  PUNTO DE ENTRADA
# ─────────────────────────────────────────────────────────────────────────────

def main():
    parser = argparse.ArgumentParser(description="ETL Landing — Fertilisa S.A. v3.0")
    parser.add_argument("--input-dir", required=True, help="Carpeta con los archivos .xlsx")
    parser.add_argument("--log-dir", default="logs", help="Carpeta para .log")
    parser.add_argument("--usuario", default=os.getenv("USER", "etl_proceso"), help="Usuario ejecutor")
    args = parser.parse_args()

    logger = configurar_logging(args.log_dir)
    logger.info("=" * 60)
    logger.info("ETL LANDING — Fertilisa S.A.")
    logger.info(f"  Script version : {SCRIPT_VERSION}")
    logger.info(f"  Carpeta entrada: {args.input_dir}")
    logger.info(f"  Usuario        : {args.usuario}")
    logger.info("=" * 60)

    carpeta = Path(args.input_dir)
    if not carpeta.exists() or not carpeta.is_dir():
        logger.critical(f"La carpeta de entrada no existe: {args.input_dir}")
        sys.exit(1)

    archivos = sorted(carpeta.glob("*.xlsx"))
    if not archivos:
        logger.warning(f"No se encontraron archivos .xlsx en: {args.input_dir}")
        sys.exit(0)

    logger.info(f"Archivos encontrados: {len(archivos)}")
    conn = conectar_db(logger)

    resultados = {"exitosos": 0, "duplicados": 0, "errores": 0}

    for ruta_archivo in archivos:
        if ruta_archivo.name.startswith("~$"):
            continue
            
        try:
            ok = procesar_archivo(ruta_archivo, conn, logger, args.usuario)
            if ok:
                resultados["exitosos"] += 1
            else:
                resultados["duplicados"] += 1
        except Exception as e:
            logger.error(f"Error inesperado procesando '{ruta_archivo.name}': {e}")
            try:
                conn.rollback()
            except Exception:
                pass
            resultados["errores"] += 1

    conn.close()

    logger.info("=" * 60)
    logger.info("RESUMEN GENERAL")
    logger.info(f"  Exitosos   : {resultados['exitosos']}")
    logger.info(f"  Duplicados : {resultados['duplicados']}")
    logger.info(f"  Errores    : {resultados['errores']}")
    logger.info("=" * 60)

    sys.exit(0 if resultados["errores"] == 0 else 1)

if __name__ == "__main__":
    main()
